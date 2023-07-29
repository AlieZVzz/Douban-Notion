import time
from notion_database.properties import Properties
from notion_database.database import Database
from notion_database.page import Page
import requests
from bs4 import BeautifulSoup
import re

import openpyxl
from requests import RequestException
from bs4 import BeautifulSoup
import lxml
import time
import random
import NotionAPI
import feedparser
from config import *







def DataBase_item_query(query_database_id):
    headers = {
    "accept": "application/json",
    "Notion-Version": "2022-06-28",
    "content-type": "application/json",
    "authorization": notion_api}

    proxies = {'http': "http://127.0.0.1:7890",
               'https': "http://127.0.0.1:7890"}
    url_notion_block = 'https://api.notion.com/v1/databases/' + query_database_id + '/query'
    res_notion = requests.post(url_notion_block, headers=headers)
    print(res_notion.json)
    S_0 = res_notion.json()
    res_travel = S_0['results']
    if_continue = len(res_travel)
    if if_continue > 0:
        while if_continue % 100 == 0:
            body = {
                'start_cursor': res_travel[-1]['id']
            }
            res_notion_plus = requests.post(url_notion_block, headers=headers, json=body, verify=False)
            S_0plus = res_notion_plus.json()
            res_travel_plus = S_0plus['results']
            for i in res_travel_plus:
                if i['id'] == res_travel[-1]['id']:
                    continue
                res_travel.append(i)
            if_continue = len(res_travel_plus)
    return res_travel



def DataBase_item_query(query_database_id):
    headers = {
    "accept": "application/json",
    "Notion-Version": "2022-06-28",
    "content-type": "application/json",
    "authorization": notion_api}

    proxies = {'http': "http://127.0.0.1:7890",
               'https': "http://127.0.0.1:7890"}
    url_notion_block = 'https://api.notion.com/v1/databases/' + query_database_id + '/query'
    res_notion = requests.post(url_notion_block, headers=headers)
    print(res_notion.json)
    S_0 = res_notion.json()
    res_travel = S_0['results']
    if_continue = len(res_travel)
    if if_continue > 0:
        while if_continue % 100 == 0:
            body = {
                'start_cursor': res_travel[-1]['id']
            }
            res_notion_plus = requests.post(url_notion_block, headers=headers, json=body, verify=False)
            S_0plus = res_notion_plus.json()
            res_travel_plus = S_0plus['results']
            for i in res_travel_plus:
                if i['id'] == res_travel[-1]['id']:
                    continue
                res_travel.append(i)
            if_continue = len(res_travel_plus)
    return res_travel



def get_one_page(url):
    '''
    Get the html of a page by requests module
    :param url: page url
    :return: html / None
    '''
    try:
        head = ['Mozilla/5.0', 'Chrome/78.0.3904.97', 'Safari/537.36']
        headers = {
            'user-agent':head[random.randint(0, 2)]
        }
        response = requests.get(url, headers=headers) #, proxies={'http':'171.15.65.195:9999'}
        if response.status_code == 200:
            return response.text
        return None
    except RequestException:
        return None

def get_request_res(pattern_text, html):
    '''
    Get the book info by re module
    :param pattern_text: re pattern
    :param html: page's html text
    :return: book's info
    '''
    pattern = re.compile(pattern_text, re.S)
    res = re.findall(pattern, html)
    if len(res) > 0:
        return res[0].split('<', 1)[0][1:]
    else:
        return 'NULL'
    

def get_request_press(pattern_text, html):
    '''
    Get the book info by re module
    :param pattern_text: re pattern
    :param html: page's html text
    :return: book's info
    '''
    pattern = re.compile(pattern_text, re.S)
    res = re.findall(pattern, html)
    if len(res) > 0:
        return res[0].split(">")[1]
        
    else:
        return 'NULL'

def get_bs_res(selector, html):
    '''
    Get the book info by bs4 module
    :param selector: info selector
    :param html: page's html text
    :return: book's info
    '''
    soup = BeautifulSoup(html, 'lxml')
    res = soup.select(selector)
    # if res is not None or len(res) is not 0:
    #     return res[0].string
    # else:
    #     return 'NULL'
    if res is None:
        return 'NULL'
    elif len(res) == 0:
        return 'NULL'
    else:
        return res[0].string

# Get other info by bs module
def get_bs_img_res(selector, html):
    soup = BeautifulSoup(html, 'lxml')
    res = soup.select(selector)
    if len(res) != 0:
        return str(res[0])
    else:
        return 'NULL'

def parse_one_page(url,headers):
    res = requests.get(url,headers=headers,allow_redirects=False)
    html = res.content.decode("utf-8")
    '''
    Parse the useful info of html by re module
    :param html: page's html text
    :return: all of book info(dict)
    '''
    book_info = {}
    book_name = get_bs_res('div > h1 > span', html)
    # print('Book-name', book_name)
    book_info['Book_name'] = book_name
    # info > a:nth-child(2)
    author = get_bs_res('div > span:nth-child(1) > a', html)
    if author is None:
        author = get_bs_res('#info > a:nth-child(2)', html)
    # print('Author', author)
    author = author.replace(" ", "")
    author = author.replace("\n", "")
    book_info['Author'] = author

    publisher = get_request_press(u'出版社:</span>([\s\S]*?)</a>', html)
    # print('Publisher', publisher)
    book_info['publisher'] = publisher

    publish_time = get_request_res(u'出版年:</span>(.*?)<br/>', html)
    # print('Publish-time', publish_time)
    book_info['publish_time'] = publish_time

    ISBN = get_request_res(u'ISBN:</span>(.*?)<br/>', html)
    # print('ISBN', ISBN)
    book_info['ISBN'] = ISBN

    img_label = get_bs_img_res('#mainpic > a > img', html)
    pattern = re.compile('src="(.*?)"', re.S)
    img = re.findall(pattern, img_label)
    if len(img) != 0:
        # print('img-src', img[0])
        book_info['img_src'] = img[0]
    else:
        # print('src not found')
        book_info['img_src'] = 'NULL'

    book_intro = get_bs_res('#link-report > div:nth-child(1) > div > p', html)
    # print('book introduction', book_intro)
    book_info['book_intro'] = book_intro

    author_intro = get_bs_res('#content > div > div.article > div.related_info > div:nth-child(4) > div > div > p', html)
    # print('author introduction', author_intro)
    book_info['author_intro'] = author_intro

    grade = get_bs_res('div > div.rating_self.clearfix > strong', html)
    if len(grade) == 1:
        # print('Score no mark')
        book_info['Score'] = 'NULL'
    else:
        # print('Score', grade[1:])
        book_info['Score'] = grade[1:]

    comment_num = get_bs_res('#interest_sectl > div > div.rating_self.clearfix > div > div.rating_sum > span > a > span', html)
    # print('commments', comment_num)
    book_info['commments'] = comment_num

    five_stars = get_bs_res('#interest_sectl > div > span:nth-child(5)', html)
    # print('5-stars', five_stars)
    book_info['5_stars'] = five_stars

    four_stars = get_bs_res('#interest_sectl > div > span:nth-child(9)', html)
    # print('4-stars', four_stars)
    book_info['4_stars'] = four_stars

    three_stars = get_bs_res('#interest_sectl > div > span:nth-child(13)', html)
    # print('3-stars', three_stars)
    book_info['3_stars'] = three_stars

    two_stars = get_bs_res('#interest_sectl > div > span:nth-child(17)', html)
    # print('2-stars', two_stars)
    book_info['2_stars'] = two_stars

    one_stars = get_bs_res('#interest_sectl > div > span:nth-child(21)', html)
    # print('1-stars', one_stars)
    book_info['1_stars'] = one_stars

    return book_info

def write_bookinfo_excel(book_info, file):
    '''
    Write book info into excel file
    :param book_info: a dict
    :param file: memory excel file
    :return: the num of successful item
    '''
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    sheet_row = ws.max_row
    sheet_col = ws.max_column
    i = sheet_row
    j = 1
    for key in book_info:
        ws.cell(i+1, j).value = book_info[key]
        j += 1
    done = ws.max_row - sheet_row
    wb.save(file)
    return done

def read_booksrc_get_info(src_file, info_file):
    '''
    Read the src file and access each src, parse html and write info into file
    :param src_file: src file
    :param info_file: memory file
    :return: the num of successful item
    '''
    wb = openpyxl.load_workbook(src_file)
    ws = wb.worksheets[0]
    row = ws.max_row
    done = 0
    for i in range(868, row+1):
        src = ws.cell(i, 1).value
        if src is None:
            continue
        html = get_one_page(str(src))
        book_info = parse_one_page(html)
        done += write_bookinfo_excel(book_info, info_file)
        if done % 10 == 0:
            print(done, 'done')
    return done




if __name__ == "__main__":
    rss_movietracker = feedparser.parse(rss_address,
                                        request_headers=douban_headers)
    notion_books = NotionAPI.DataBase_item_query(book_database)
    # print(notion_books)
    books = [item['properties']['链接']['url'] for item in notion_books]
    print(rss_movietracker)
    for item in rss_movietracker["entries"]:
        print(item)
        # if "看过" not in item["title"]:
        #     continue

        # rel = NotionAPI.select_items_form_Databaseitems(notion_books, "链接", movie_url)
        if "读过" in item["title"]:
            title = item["title"].split("读过")[1]
        elif "最近在读" in item["title"]:
            title = item["title"].split("最近在读")[1]
        else:
            continue
        time1 = item["published"]
        pattern1 = re.compile(r'(?<=src=").+(?=")', re.I)  
        cover_url = re.findall(pattern1, item["summary"])[0]
        book_url = item["link"]
        cover_url = cover_url.replace("s_ratio_poster", "r")
        pattern2 = re.compile(r'(?<=. ).+\d{4}', re.S)  # 匹配时间
        month_satandard = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
                        'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
        time1 = re.findall(pattern2, time1)[0]
        time1 = time1.split(" ")
        day = time1[0]
        month = month_satandard[time1[1]]
        year = time1[2]
        watch_time = str(year) + "-" + str(month) + "-" + str(day)
        pattern = re.compile(r'(?<=<p>).+(?=</p>)', re.S)  # 匹配评论·
        # pattern2 = re.compile(r'(?<=<p>)(.|\n)+(?=</p>)', re.I) # 匹配评论·
        allcomment = re.findall(pattern, item["summary"])[0]  # 需要进一步处理
        # print(allcomment)

        pattern1 = re.compile(r'(?<=推荐: ).+(?=</p>)', re.S)  # 匹配评分
        # 一星：很差 二星：较差 三星：还行 四星：推荐 五星：力荐
        scoredict = {'很差': '⭐', '较差': '⭐⭐', '还行': '⭐⭐⭐', '推荐': '⭐⭐⭐⭐', '力荐': '⭐⭐⭐⭐⭐', }
        # score = re.findall(pattern1, allcomment)
        score = allcomment[-2:]
        comment = ""
        if score:
            score = scoredict[score]
        else:
            score = "⭐⭐⭐"

        if book_url not in books:
            book_info = parse_one_page(book_url, headers=douban_headers)
            body = {
                        'properties': {
                            '标题': {
                                'title': [{'type': 'text', 'text': {'content': str(title)}}]
                            },
                            '阅读时间': {'date': {'start': str(watch_time)}},
                            '评分': {'type': 'select', 'select': {'name': str(score)}},
                            'Author': {'type': 'select', 'select': {'name': str(book_info["Author"])}},
                            'Publisher': {'type': 'select', 'select': {'name': str(book_info["publisher"])}},
                            'Cover': {
                                'files': [{'type': 'external', 'name': '封面', 'external': {'url': str(cover_url)}}]
                            },
                            '链接': {'type': 'url', 'url': str(book_url)},
                            'ISBN': 
                                {'type': 'rich_text', 'rich_text': [{'type': 'text', 'text': {'content': str(book_info["ISBN"])},  'plain_text': str(book_info["ISBN"])}]},
                            
                            }

                        }
            print(body)
            NotionAPI.DataBase_additem(book_database, body, title)
            time.sleep(3)
    with open("start.txt") as f:
        print(f.read())                    
        